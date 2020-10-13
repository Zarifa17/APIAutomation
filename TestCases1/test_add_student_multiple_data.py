import requests
import json
import jsonpath
import openpyxl

def test_add_mulitple_data():
    #API
    API_URL = "http://thetestingworldapi.com/api/studentsDetails"
    file = open("/Users/farhanmohammed/PycharmProjects/pythonPyTest/TestCases1/RequestJson.json",'r')
    json_request = json.loads(file.read())

    #Excel Code
    wk = openpyxl.load_workbook('/Users/farhanmohammed/PycharmProjects/pythonPyTest/TestCases1/TestData.xlsx') # Create workbook object
    sh = wk['Sheet1'] # Create sheet object
    rows = sh.max_row

    for i in range(2,rows+1):
        # this reads data from each cell in the excel sheet
        cell_first_name = sh.cell(row=i, column=1)
        cell_mid_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_dob = sh.cell(row=i, column=4)

        # Now it puts/assigns this data in the json request. It updates the value of the keys.
        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_mid_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_dob.value


        response = requests.post(API_URL,json_request)

        print(response.text)
        print(response.status_code)
        assert response.status_code==201
